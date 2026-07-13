"""Configurable importer for NAPIC residential count/value workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path
import re
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from ..data_pipeline import aggregate_property_type


@dataclass(frozen=True)
class NapicWorkbookLayout:
    """Layout rules shared by compatible state workbooks."""

    count_title: str = "Breakdown Of Number Of Residential Property Transactions"
    value_title: str = "Breakdown Of Value Of Residential Property Transactions"
    excluded_property_types: tuple[str, ...] = ("Vacant Plot", "Others", "Total")
    value_multiplier: float = 1_000_000.0


class NapicExcelImporter:
    """Join residential count/value sheets without state-specific code paths."""

    def __init__(self, layout: NapicWorkbookLayout | None = None) -> None:
        self.layout = layout or NapicWorkbookLayout()

    def import_file(
        self,
        path: str | Path,
        *,
        state: str,
        source_url: str,
        retrieved_at: str,
        dataset_version: str,
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        source = Path(path)
        workbook = load_workbook(source, data_only=True, read_only=False)
        count_sheet = self._find_sheet(workbook, self.layout.count_title)
        value_sheet = self._find_sheet(workbook, self.layout.value_title)
        count_rows = self._table_rows(count_sheet)
        value_rows = self._table_rows(value_sheet)
        if len(count_rows) != len(value_rows):
            raise ValueError("NAPIC count/value sheets have different row counts")

        accepted: list[dict[str, Any]] = []
        rejected: list[dict[str, Any]] = []
        for count, value in zip(count_rows, value_rows, strict=True):
            coordinates = {
                "state": state,
                "source_file": source.name,
                "source_sheet": f"{count_sheet.title}|{value_sheet.title}",
                "source_row": f"{count['row']}|{value['row']}",
            }
            errors: list[str] = []
            if count["property_type"] != value["property_type"]:
                errors.append("PROPERTY_TYPE_MISMATCH")
            if count["period"] != value["period"]:
                errors.append("PERIOD_MISMATCH")
            if tuple(count["districts"]) != tuple(value["districts"]):
                errors.append("DISTRICT_LAYOUT_MISMATCH")
            if errors:
                rejected.append({**coordinates, "validation_errors": errors})
                continue
            if count["property_type"] in self.layout.excluded_property_types:
                continue
            period = self._parse_period(count["period"])
            if period is None:
                rejected.append({**coordinates, "validation_errors": ["UNRECOGNISED_PERIOD"]})
                continue
            year, quarter = period
            property_definition = aggregate_property_type(count["property_type"])
            for district, raw_count in count["districts"].items():
                raw_value = value["districts"].get(district)
                transaction_count = self._positive_int(raw_count)
                transaction_value = self._positive_float(raw_value)
                if transaction_count is None or transaction_value is None:
                    if raw_count not in (None, 0, "-") or raw_value not in (None, 0, "-"):
                        rejected.append(
                            {
                                **coordinates,
                                "district": district,
                                "property_type": count["property_type"],
                                "period": count["period"],
                                "validation_errors": ["INVALID_COUNT_OR_VALUE"],
                            }
                        )
                    continue
                value_rm = transaction_value * self.layout.value_multiplier
                period_value = pd.Period(year=year, quarter=quarter, freq="Q")
                accepted.append(
                    {
                        "state": state,
                        "district": district,
                        "district_notes": "",
                        "property_type": property_definition.storage_code,
                        "year": year,
                        "period_type": "quarter",
                        "period_number": quarter,
                        "period_start": period_value.start_time.date().isoformat(),
                        "period_end": period_value.end_time.date().isoformat(),
                        "transaction_count": transaction_count,
                        "transaction_value_rm": value_rm,
                        "average_price_rm": value_rm / transaction_count,
                        "price_type": "completed_transaction_average",
                        "source_name": "NAPIC/JPPH Property Transaction Data Tables",
                        "source_dataset": "NAPIC state transaction publication",
                        "source_file": source.name,
                        "source_sheet": coordinates["source_sheet"],
                        "source_row": coordinates["source_row"],
                        "source_url": source_url,
                        "source_document": source.stem,
                        "source_table": "Residential count and value by type and district",
                        "retrieved_at": retrieved_at,
                        "dataset_version": dataset_version,
                        "schema_version": "aggregate-2.0.0",
                        "validation_status": "valid",
                        "validation_errors": "[]",
                        "validation_warnings": (
                            "[]" if property_definition.known
                            else '["UNKNOWN_PROPERTY_TYPE"]'
                        ),
                    }
                )
        return pd.DataFrame(accepted), pd.DataFrame(rejected)

    def _find_sheet(self, workbook: Any, title_fragment: str) -> Any:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=min(4, sheet.max_row), values_only=True):
                if any(title_fragment.lower() in str(value).lower() for value in row if value):
                    return sheet
        raise ValueError(f"Unrecognised NAPIC layout: missing sheet titled {title_fragment!r}")

    @staticmethod
    def _header_row(sheet: Any) -> int:
        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [str(value).strip().lower() if value is not None else "" for value in row]
            if "property type" in values and "quarter" in values:
                return row_number
        raise ValueError(f"Unrecognised NAPIC layout in sheet {sheet.title}: header not found")

    def _table_rows(self, sheet: Any) -> list[dict[str, Any]]:
        header_row = self._header_row(sheet)
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[header_row]]
        district_indexes = {
            index: header
            for index, header in enumerate(headers[2:], start=2)
            if header and not header.lower().startswith("total")
        }
        period_cycle: list[str] = []
        for row_number in range(header_row + 1, sheet.max_row + 1):
            value = sheet.cell(row_number, 2).value
            label = str(value).strip() if value is not None else ""
            if self._parse_period(label) and label not in period_cycle:
                period_cycle.append(label)
        if not period_cycle:
            raise ValueError(f"Unrecognised NAPIC layout in sheet {sheet.title}: periods not found")
        rows: list[dict[str, Any]] = []
        property_type: str | None = None
        cycle_index = 0
        for row_number in range(header_row + 1, sheet.max_row + 1):
            values = [sheet.cell(row_number, column + 1).value for column in range(len(headers))]
            if not any(value is not None for value in values):
                continue
            if values[0] is not None:
                property_type = str(values[0]).strip()
                cycle_index = 0
            explicit_period = str(values[1]).strip() if values[1] is not None else ""
            if not property_type:
                continue
            period = explicit_period or period_cycle[cycle_index % len(period_cycle)]
            rows.append(
                {
                    "row": row_number,
                    "property_type": property_type,
                    "period": period,
                    "districts": {name: values[index] for index, name in district_indexes.items()},
                }
            )
            cycle_index += 1
            if cycle_index == len(period_cycle):
                cycle_index = 0
        return rows

    @staticmethod
    def _parse_period(value: str) -> tuple[int, int] | None:
        match = re.search(r"Q([1-4])\s+(20\d{2})", value, flags=re.IGNORECASE)
        return (int(match.group(2)), int(match.group(1))) if match else None

    @staticmethod
    def _slug(value: str) -> str:
        return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")

    @staticmethod
    def _positive_int(value: object) -> int | None:
        try:
            parsed = float(value)
        except (TypeError, ValueError):
            return None
        return int(parsed) if parsed > 0 and parsed.is_integer() else None

    @staticmethod
    def _positive_float(value: object) -> float | None:
        try:
            parsed = float(value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None
